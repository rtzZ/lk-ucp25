"""Синхронизация данных из публичных Яндекс.Таблиц (disk.360).

Цепочка: edit-страница рендерится headless Chromium (иначе прямых ссылок
в HTML нет) -> из HTML извлекается подписанный URL downloader.disk.yandex.ru
(подпись покрывает query целиком, менять параметры нельзя) -> XLSX качается
обычным httpx -> парсинг openpyxl -> полная замена данных в БД.

Структура файлов (установлена по реальным таблицам УЦП-25):
- Расписание: листы-семестры + листы других групп; шапка: дни недели (строка 1,
  merged) + время пар (строка 2, «HH:MM-HH:MM»); далее блоки недель: строка
  с датами + строки пар. Ячейка пары — многострочный текст:
  «Предмет\\nОрганизация\\nНомер\\nПреподаватель» плюс маркеры ОТМЕНА /
  ПЕРЕНОС / ДЕДЛАЙН / события. Колонки правее воскресенья — аттестации,
  в расписание не идут.
- Успеваемость: листы-семестры; строка 5 (merged) — блоки предметов
  («Тип аттестации» слева + название), строка 7 — задания («Итоговая оценка»,
  ECTS...), строка 8 — максимумы, строка 9 — дедлайны, далее строки студентов.
  ФИО в блоках короткие (без отчества) — матч по «фамилия+имя».

При недоступности ссылок ошибка логируется, БД хранит последние данные.
"""

import datetime
import io
import os
import re

import httpx
from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import delete, func, select

from .db import SessionLocal
from .models import Grade, Group, ScheduleItem, Student, Subject

SCHEDULE_EDIT_URL = os.getenv("YANDEX_SCHEDULE_URL", "")
GRADES_EDIT_URL = os.getenv("YANDEX_GRADES_URL", "")

DOWNLOADER_RE = re.compile(r"http://localhost:12701/disk/[^\"'\\\s<>]+")
TIME_RE = re.compile(r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})")
PERSON_RE = re.compile(
    r"^[А-ЯЁA-Z][а-яёa-z]+\s+[А-ЯЁA-Z][а-яёa-z.\-]+"
    r"(?:\s+[А-ЯЁA-Z][а-яёa-z.\-]+)?$"
)
ORG_RE = re.compile(r"ранхигс|нетология|\bсдо\b", re.IGNORECASE)
LINK_RE = re.compile(r"https?://[^\s<>\"\u00ab\u00bb]+")
ATTESTATION_TYPES = {"экзамен", "зачет", "зачёт", "зачет с оценкой", "зачёт с оценкой"}
# Шкалы ведомости: итоговая 1-5 <-> словесная, баллы 0-100 <-> ECTS.
VERBAL_GRADES = {"отлично", "хорошо", "удовлетворительно", "удов-но",
                 "удовл", "удовлетв", "неудовлетворительно", "неуд"}
ECTS_RE = re.compile(r"^(a|b|c|d|e|f|passed)$", re.IGNORECASE)
SCORE_HEADERS = {"всего баллов", "итого", "итого баллов"}
WEEKDAYS = {
    "пн": 1, "понедельник": 1, "1": 1,
    "вт": 2, "вторник": 2, "2": 2,
    "ср": 3, "среда": 3, "среду": 3, "3": 3,
    "чт": 4, "четверг": 4, "4": 4,
    "пт": 5, "пятница": 5, "пятницу": 5, "5": 5,
    "сб": 6, "суббота": 6, "субботу": 6, "6": 6,
    "вс": 7, "воскресенье": 7, "7": 7,
}
WEEK_MARKERS = {"сессия", "каникулы"}


def parse_weekday(value: str) -> int:
    """'Пн'/'понедельник'/'1' -> 1..7, иначе ValueError."""
    key = str(value).strip().lower()
    if key not in WEEKDAYS:
        raise ValueError(f"Неизвестный день недели: {value!r}")
    return WEEKDAYS[key]


def parse_time_range(value: str) -> tuple[str, str] | None:
    """'19:00-20:20' -> ('19:00', '20:20'), иначе None."""
    m = TIME_RE.search(str(value or ""))
    return (m.group(1), m.group(2)) if m else None


def norm_name(value: str) -> str:
    """'  Аксёнов   Роман ' -> 'аксёнов роман'."""
    return " ".join(str(value or "").split()).lower()


def split_fio(full: str) -> tuple[str, str, str]:
    """'Азимов Мурад Анар Оглы' -> ('Азимов', 'Мурад', 'Азимов Мурад Анар Оглы')."""
    parts = " ".join(str(full or "").split()).split(" ")
    last = parts[0] if len(parts) > 0 else ""
    first = parts[1] if len(parts) > 1 else ""
    return last, first, " ".join(parts)


def student_code(last: str, first: str) -> str:
    """Стабильный код студента: нормализованные 'фамилия имя'."""
    return norm_name(f"{last} {first}")


def parse_date(value) -> datetime.date | None:
    """datetime/date/строка ('01.09.2026', '2026-09-01', '07.06;2026') -> date."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip().replace(";", ".")
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


TITLE_PAREN_RE = re.compile(
    r"\((зач[её]т(?:\s+с\s+оценкой)?|экзамен)\)", re.IGNORECASE)
TITLE_TAIL_RE = re.compile(
    r"\s+(зач[её]т(?:\s+с\s+оценкой)?|экзамен)\s*$", re.IGNORECASE)
TITLE_ORG_PAREN_RE = re.compile(
    r"\((?:ранхигс|нетология|сдо)[^)]*\)", re.IGNORECASE)
FORMULA_RE = re.compile(r"^(итоговая|кт\s*\d|успеваемость)", re.IGNORECASE)


def split_subject_title(title: str) -> tuple[str, str]:
    """Шапка блока -> (название предмета, тип аттестации из скобок).

    Убирает организацию, преподавателя и «(зачет с оценкой)»-пометки.
    """
    attest = ""
    m = TITLE_PAREN_RE.search(str(title or ""))
    if m:
        attest = norm_name(m.group(1)).replace("ё", "е")
        title = TITLE_PAREN_RE.sub("", str(title))
    else:
        m = TITLE_TAIL_RE.search(str(title or ""))
        if m:
            attest = norm_name(m.group(1)).replace("ё", "е")
            title = TITLE_TAIL_RE.sub("", str(title))
    # Организация в скобках внутри строки названия — вырезать, строку оставить.
    title = TITLE_ORG_PAREN_RE.sub("", str(title))
    keep = [
        line.strip()
        for line in str(title or "").split("\n")
        if line.strip()
        and not ORG_RE.search(line)
        and not PERSON_RE.match(line.strip())
    ]
    return " ".join(keep), attest


def clean_subject_title(title: str) -> str:
    """Название предмета из шапки блока (без типа аттестации)."""
    return split_subject_title(title)[0]


def _is_formula_text(text: str) -> bool:
    """Строка формул/подписей, а не название предмета."""
    t = str(text or "").strip()
    return not t or bool(FORMULA_RE.match(t)) or "=" in t.split("\n")[0]


def parse_lesson_cell(text: str) -> dict | None:
    """Ячейка пары -> dict(kind, subject, teacher, org, lesson_no, status, note).

    None — служебная ячейка (Сессия/Каникулы/пусто).
    kind: lesson | attestation | event | deadline.
    status: active | cancelled | moved.
    """
    lines = [ln.strip() for ln in str(text or "").split("\n")]
    lines = [ln for ln in lines if ln]
    if not lines:
        return None
    low = [ln.lower() for ln in lines]
    if len(lines) == 1 and lines[0].lower() in WEEK_MARKERS:
        return None

    cancelled = any("отмена" in ln for ln in low)
    moved = any("перенос" in ln for ln in low)
    status = "cancelled" if cancelled else ("moved" if moved else "active")

    marker_lines = (
        "отмена", "перенос", "дедлайн", "дедлаин", "deadline",
        "сессия", "каникулы", "пересдача",
        "зачет", "зачёт", "экзамен", "асинхрон", "защита",
    )
    content = [ln for ln in lines if not any(m in ln.lower() for m in marker_lines)]

    blob = " ".join(low)
    if any(m in blob for m in ("дедлайн", "дедлаин", "deadline")):
        kind = "deadline"
    elif "встреча" in blob:
        kind = "event"
    elif any(k in blob for k in ("зачет", "зачёт", "экзамен", "защита")):
        kind = "attestation"
    else:
        kind = "lesson"

    subject = next((ln for ln in content if ln and not ln.isdigit()), "")
    if kind == "lesson" and not subject:
        return None
    if not subject:
        subject = lines[0][:512]  # одиночные «Дедлайн ...» / «зачет, ...»
    org_m = next((ln for ln in content if ORG_RE.search(ln)), "")
    org = "РАНХиГС" if "ранхигс" in org_m.lower() else (
        "Нетология" if "нетология" in org_m.lower() else org_m)
    teacher = next(
        (ln for ln in content
         if ln != subject and PERSON_RE.match(ln) and not ORG_RE.search(ln)),
        "",
    )
    no_m = next((ln for ln in content if re.fullmatch(r"\d{1,3}", ln)), None)
    link_m = next((LINK_RE.search(ln) for ln in lines if LINK_RE.search(ln)),
                  None)

    return {
        "kind": kind,
        "subject": subject,
        "teacher": teacher,
        "org": org,
        "lesson_no": int(no_m) if no_m else None,
        "status": status,
        "note": "" if kind == "lesson" and status == "active"
        else " / ".join(lines)[:1024],
        "link": link_m.group(0) if link_m else "",
    }


def extract_downloader_url(edit_url: str) -> str:
    """Прямая ссылка на XLSX из отрендеренной edit-страницы."""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page()
            page.goto(edit_url, wait_until="domcontentloaded", timeout=90000)
            page.wait_for_timeout(12000)
            html = page.content().replace("&amp;", "&")
        finally:
            browser.close()
    urls = DOWNLOADER_RE.findall(html)
    xlsx = [u for u in urls if "spreadsheetml" in u]
    if not xlsx:
        raise RuntimeError("Прямая ссылка на файл не найдена в HTML страницы")
    return xlsx[0].replace("http://localhost:12701",
                           "https://downloader.disk.yandex.ru")


async def download_xlsx(url: str) -> tuple[bytes, str]:
    """Скачивание XLSX. Возвращает (байты, имя файла)."""
    async with httpx.AsyncClient(timeout=120, follow_redirects=True) as client:
        resp = await client.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        })
        resp.raise_for_status()
        name = httpx.QueryParams(httpx.URL(url).query).get("filename", "file.xlsx")
        return resp.content, name


def _merged_values(ws) -> dict[tuple[int, int], object]:
    """Значение левого верхнего угла merged-диапазона для каждой его ячейки."""
    filled: dict[tuple[int, int], object] = {}
    for mr in ws.merged_cells.ranges:
        value = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                filled[(r, c)] = value
    return filled


def _sheet_group(sheet_name: str, file_group: str) -> str | None:
    """Группа листа: листы «N семестр» — группа файла, чужие листы — None.

    В файле расписания встречаются листы других групп (напр. «ЛФТ-25 2 сем»)
    — они пропускаются: в кабинете только своя группа.
    """
    if re.match(r"^\d+\s*семестр", sheet_name.strip(), re.IGNORECASE):
        return file_group
    return None


def _file_group(filename: str, prefix: str) -> str:
    """'Расписание УЦП-25 1 сем...' -> 'УЦП-25'."""
    m = re.search(re.escape(prefix) + r"\s+(.+?)(?:\s+\d|\.|$)", filename,
                  re.IGNORECASE)
    return m.group(1).strip() if m else "УЦП-25"


def parse_schedule_xlsx(payload: bytes, filename: str) -> list[dict]:
    """XLSX расписания -> список занятий (dict с group/date/time/...)."""
    wb = load_workbook(io.BytesIO(payload), data_only=True)
    file_group = _file_group(filename, "Расписание")
    items: list[dict] = []
    for sheet in wb.worksheets:
        group = _sheet_group(sheet.title, file_group)
        if group is None:
            logger.warning(f"[{sheet.title}] чужой лист, пропуск")
            continue
        merged = _merged_values(sheet)

        def val(r, c):
            return merged.get((r, c), sheet.cell(r, c).value)

        # Шапка: дни недели (строка 1) + время пар (строка 2).
        day_of: dict[int, int] = {}
        for c in range(2, sheet.max_column + 1):
            try:
                day_of[c] = parse_weekday(val(1, c) or "")
            except ValueError:
                continue
        if not day_of:
            logger.warning(f"[{sheet.title}] шапка дней не распознана, пропуск")
            continue
        slot_of: dict[int, tuple[str, str] | None] = {}
        for c, day in day_of.items():
            slot_of[c] = parse_time_range(val(2, c) or "")
        lesson_cols = [c for c in day_of
                       if slot_of[c] is not None or day_of[c] == 7]
        if not lesson_cols:
            continue

        dates: dict[int, datetime.date] = {}
        for r in range(3, sheet.max_row + 1):
            row_dates = {c: parse_date(val(r, c)) for c in lesson_cols}
            if any(row_dates.values()):
                dates = {c: d for c, d in row_dates.items() if d}
                continue
            for c in lesson_cols:
                cell = val(r, c)
                if not isinstance(cell, str) or not cell.strip():
                    continue  # даты/время/числа — не пары
                if c not in dates:
                    continue
                lesson = parse_lesson_cell(cell)
                if lesson is None:
                    continue
                start, end = slot_of[c] or ("", "")
                items.append({
                    "group": group,
                    "date": dates[c].isoformat(),
                    "time_start": start, "time_end": end,
                    "subject_text": lesson["subject"],
                    "teacher": lesson["teacher"],
                    "org": lesson["org"],
                    "lesson_no": lesson["lesson_no"],
                    "kind": lesson["kind"],
                    "status": lesson["status"],
                    "note": lesson["note"],
                })
    return items


def parse_grades_xlsx(payload: bytes) -> tuple[list[dict], list[dict]]:
    """XLSX успеваемости -> (студенты, оценки).

    Студент: {group, code, last_name, first_name, full_name}.
    Оценка: {student_code, subject, semester, attestation, value, verbal,
    ects, score}.
    """
    wb = load_workbook(io.BytesIO(payload), data_only=True)
    students: dict[str, dict] = {}
    grades: list[dict] = []
    group = "УЦП-25"  # группа из имени файла «Успеваемость сводная УЦП-25»

    for sheet in wb.worksheets:
        if "семестр" not in sheet.title.lower():
            continue  # листы команд — не оценки
        semester = " ".join(sheet.title.split())
        merged = _merged_values(sheet)

        def val(r, c):
            return merged.get((r, c), sheet.cell(r, c).value)

        # Строка-заголовок заданий: ищем ряд с 'ECTS'.
        header_row = next(
            (r for r in range(1, 15)
             if any(str(val(r, c) or "").strip().upper() == "ECTS"
                    for c in range(1, sheet.max_column + 1))),
            None,
        )
        if header_row is None:
            logger.warning(f"[{sheet.title}] строка заданий не найдена, пропуск")
            continue

        # Блоки предметов из merged-шапок над строкой заданий.
        # Название может быть в строке header-1 (2 семестр) или header-2
        # (1 семестр); строки формул пропускаются.
        blocks: list[tuple[int, int, str, str]] = []  # min_c, max_c, subject, attest
        seen: set[tuple[int, int]] = set()
        for title_row in (header_row - 1, header_row - 2):
            if title_row < 1:
                continue
            for mr in sheet.merged_cells.ranges:
                if not (mr.min_row <= title_row <= mr.max_row):
                    continue
                if mr.max_col <= mr.min_col:
                    continue
                key = (mr.min_col, mr.max_col)
                if key in seen:
                    continue
                raw_title = val(title_row, mr.min_col) or ""
                if _is_formula_text(raw_title):
                    continue
                subject, attest_par = split_subject_title(raw_title)
                if not subject:
                    continue
                seen.add(key)
                attest = norm_name(val(title_row, mr.min_col - 1) or "")
                if attest not in ATTESTATION_TYPES:
                    attest = next(
                        (norm_name(val(header_row, c) or "")
                         for c in range(max(mr.min_col - 3, 1), mr.min_col)
                         if norm_name(val(header_row, c) or "")
                         in ATTESTATION_TYPES),
                        "",
                    )
                if attest not in ATTESTATION_TYPES:
                    attest = attest_par
                blocks.append((mr.min_col, mr.max_col, subject, attest))
        if not blocks:
            continue

        # Основной список группы — колонка B ниже заголовка.
        first_student_row = header_row + 3
        roster_by_row: dict[int, str] = {}
        for r in range(first_student_row, sheet.max_row + 1):
            cell_b = val(r, 2)
            if not (isinstance(cell_b, str) and len(cell_b.split()) >= 2):
                continue
            if norm_name(cell_b) in WEEK_MARKERS:
                continue
            last, first, full = split_fio(cell_b)
            code = student_code(last, first)
            roster_by_row[r] = code
            if code not in students:
                students[code] = {"group": group, "code": code,
                                  "last_name": last, "first_name": first,
                                  "full_name": full}

        # Оценки по блокам. Якорь — колонка баллов («Итого»/«Всего»):
        # [баллы] [итоговая 1-5/зачтено] [словесная] [ECTS].
        # Шапки врут («Итоговая защита», пустые, ECTS над словесной),
        # поэтому словесная/ECTS определяются по значениям, не по шапкам.
        for min_c, max_c, subject, attest in blocks:
            headers = {c: norm_name(val(header_row, c) or "")
                       for c in range(min_c, max_c + 1)}
            score_c = next((c for c, h in headers.items()
                            if h in SCORE_HEADERS), None)
            if score_c is None:
                # Префиксный поиск с конца; «итоговая ...» — не баллы.
                score_c = next(
                    (c for c in sorted(headers, reverse=True)
                     if (headers[c].startswith("всего")
                         or (headers[c].startswith("итого")
                             and not headers[c].startswith("итоговая")))),
                    None)
            if score_c is not None:
                value_c = next(
                    (c for c in range(score_c + 1, min(score_c + 4, max_c + 1))
                     if not headers[c]
                     or headers[c].startswith("итоговая оценка")),
                    None)
            else:
                # Без колонки баллов — по шапке «Итоговая оценка» (не «защита»).
                value_c = next(
                    (c for c, h in headers.items()
                     if h.startswith("итоговая оценка")),
                    None)
            if value_c is None:
                continue
            verbal_c, ects_c = _verbal_ects_cols(
                val, sheet.max_row, first_student_row, value_c)
            fio_c = _block_name_col(val, sheet.max_row, first_student_row,
                                    min_c, value_c)
            # Свой список у блока (факультатив) или общий (B)?
            positional = (fio_c is not None
                          and not _is_auxiliary_list(
                              val, sheet.max_row, first_student_row,
                              fio_c, _roster_keys(students)))
            for r in range(first_student_row, sheet.max_row + 1):
                if positional:
                    candidate = str(val(r, fio_c) or "")
                    if not candidate.strip():
                        continue
                    code = _match_student(candidate, students)
                    if code is None:
                        continue
                else:
                    code = roster_by_row.get(r)
                    if code is None:
                        continue
                raw_value = val(r, value_c)
                if raw_value in (None, ""):
                    # Пустая ячейка = минимальная оценка. Свои списки
                    # (факультатив) затрагивают только своих участников.
                    minimal = _minimal_grade(attest)
                    value = minimal["value"]
                    verbal = minimal["verbal"]
                    ects = minimal["ects"]
                    score = minimal["score"]
                else:
                    value = _s(raw_value)
                    verbal = _s(val(r, verbal_c)) if verbal_c else ""
                    ects = _s(val(r, ects_c)) if ects_c else ""
                    raw_score = val(r, score_c) if score_c else None
                    score = (float(raw_score)
                             if isinstance(raw_score, (int, float)) else None)
                grades.append({
                    "student_code": code,
                    "subject": subject,
                    "semester": semester,
                    "attestation": attest,
                    "value": value,
                    "verbal": verbal,
                    "ects": ects,
                    "score": score,
                })
    return list(students.values()), grades


def _s(value) -> str:
    """Ячейка -> строка: None и мусор ('None', '-') -> пусто."""
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"none", "null", "nan", "-", "—"} else text


def _minimal_grade(attest: str) -> dict:
    """Минимальная оценка за пустую ячейку: 5-балльная -> '2',
    зачётная -> 'не зачтено'."""
    if attest in ("зачет", "зачёт"):
        return {"value": "не зачтено", "verbal": "", "ects": "F",
                "score": 0.0}
    return {"value": "2", "verbal": "Неудовлетворительно", "ects": "F",
            "score": 0.0}


def _col_values(val, max_row, first_row, c) -> list[str]:
    """Непустые значения колонки, нормализованные."""
    out = []
    for r in range(first_row, max_row + 1):
        v = val(r, c)
        if v not in (None, ""):
            out.append(norm_name(str(v)).replace("ё", "е"))
    return out


def _verbal_ects_cols(val, max_row, first_row,
                      value_c) -> tuple[int | None, int | None]:
    """Колонки словесной оценки и ECTS после итоговой — по значениям.

    Шапкам верить нельзя (пустые, «ECTS» над словесной), шкалы же строгие:
    словесная — отлично/хорошо/..., ECTS — A-E/passed.
    """
    verbal_c, ects_c = None, None
    for c in (value_c + 1, value_c + 2):
        vals = _col_values(val, max_row, first_row, c)
        if not vals:
            continue
        if verbal_c is None and all(v in VERBAL_GRADES for v in vals):
            verbal_c = c
            continue
        if ects_c is None and any(ECTS_RE.match(v) for v in vals):
            ects_c = c
    return verbal_c, ects_c


def _person_key(name: str) -> tuple[str, str]:
    """Ключ matching'а: (фамилия, имя), ё->е (отчество игнорируется)."""
    last, first, _ = split_fio(name)
    return (norm_name(last).replace("ё", "е"),
            norm_name(first).replace("ё", "е"))


def _roster_keys(students: dict[str, dict]) -> set[tuple[str, str]]:
    """Ключи (фамилия, имя) основного списка студентов."""
    return {_person_key(s["full_name"]) for s in students.values()}


def _is_auxiliary_list(val, max_row, first_row, fio_c,
                       roster_keys: set[tuple[str, str]]) -> bool:
    """Колонка имён — вспомогательная (команды), а не свой список блока.

    Вспомогательная почти полностью пересекается с основным списком
    (совпадение по фамилии+имени >= 90%), свой список (факультатив) — нет.
    """
    names = [_person_key(str(val(r, fio_c) or ""))
             for r in range(first_row, max_row + 1)]
    names = [n for n in names if n[0]]
    if not names:
        return True
    matched = sum(1 for n in names if n in roster_keys)
    return matched / len(names) >= 0.9


def _match_student(name: str, students: dict[str, dict]) -> str | None:
    """Код студента по ФИО из блока: точно -> нечётко (>=0.9) -> None."""
    import difflib

    key = _person_key(name)
    if not key[0]:
        return None
    codes = {c: _person_key(s["full_name"]) for c, s in students.items()}
    for code, k in codes.items():
        if k == key:
            return code
    swapped = (key[1], key[0])  # в ведомости бывает «Имя Фамилия»
    for code, k in codes.items():
        if k == swapped and key[1]:
            logger.warning(f"Перевёрнутое ФИО в ведомости: {name!r}")
            return code
    best, best_ratio = None, 0.0
    target = " ".join(key)
    for code, k in codes.items():
        ratio = difflib.SequenceMatcher(None, target, " ".join(k)).ratio()
        if ratio > best_ratio:
            best, best_ratio = code, ratio
    if best_ratio >= 0.9:
        return best
    logger.warning(f"Студент из ведомости не найден в списке: {name!r}")
    return None


def _block_name_col(val, max_row, first_row, min_c, final_c) -> int | None:
    """Колонка имён левее итога блока (свой список / вспомогательная)."""
    best, best_hits = None, 0
    for c in range(max(min_c - 3, 1), final_c):
        hits = sum(1 for r in range(first_row, max_row + 1)
                   if isinstance(val(r, c), str)
                   and len(str(val(r, c)).split()) >= 2)
        if hits > best_hits:
            best, best_hits = c, hits
    return best if best_hits >= 3 else None


async def _get_or_create(session, model, **kwargs):
    obj = (await session.execute(select(model).filter_by(**kwargs))
           ).scalar_one_or_none()
    if obj is None:
        obj = model(**kwargs)
        session.add(obj)
        await session.flush()
    return obj


async def upsert_schedule(items: list[dict], session) -> int:
    """Полная замена расписания. Возвращает число записей."""
    await session.execute(delete(ScheduleItem))
    groups: dict[str, Group] = {}
    for it in items:
        if it["group"] not in groups:
            groups[it["group"]] = await _get_or_create(
                session, Group, name=it["group"])
        session.add(ScheduleItem(group_id=groups[it["group"]].id, **{
            k: v for k, v in it.items() if k != "group"}))
    await session.flush()
    await _delete_orphan_groups(session)
    await session.commit()
    return len(items)


async def _delete_orphan_groups(session) -> None:
    """Группы без студентов и пар (напр. чужие из старых синхронизаций)."""
    orphan_ids = (
        select(Group.id)
        .outerjoin(Student, Student.group_id == Group.id)
        .outerjoin(ScheduleItem, ScheduleItem.group_id == Group.id)
        .group_by(Group.id)
        .having(func.count(Student.id) == 0)
        .having(func.count(ScheduleItem.id) == 0)
    )
    await session.execute(delete(Group).where(Group.id.in_(orphan_ids)))


async def upsert_grades(students: list[dict], grades: list[dict], session) -> int:
    """Полная замена студентов и оценок. Возвращает число оценок."""
    await session.execute(delete(Grade))
    await session.execute(delete(Student))
    group = await _get_or_create(session, Group, name="УЦП-25")
    by_code: dict[str, Student] = {}
    for s in students:
        g = group if s["group"] == "УЦП-25" else await _get_or_create(
            session, Group, name=s["group"])
        st = Student(group_id=g.id, **{k: v for k, v in s.items()
                                       if k != "group"})
        session.add(st)
        by_code[s["code"]] = st
    await session.flush()
    subjects: dict[str, Subject] = {}
    for gr in grades:
        if gr["subject"] not in subjects:
            subjects[gr["subject"]] = await _get_or_create(
                session, Subject, name=gr["subject"])
        st = by_code.get(gr["student_code"])
        if st is None:
            continue
        session.add(Grade(student_id=st.id, subject_id=subjects[gr["subject"]].id,
                          semester=gr["semester"], attestation=gr["attestation"],
                          value=gr["value"], verbal=gr["verbal"],
                          ects=gr["ects"], score=gr["score"]))
    await session.commit()
    return len(grades)


async def sync_source(name: str, edit_url: str, kind: str) -> None:
    """Синхронизация одного источника. Ошибки логируются, БД не трогается."""
    if not edit_url:
        logger.warning(f"[{name}] ссылка не задана, пропуск")
        return
    try:
        import asyncio
        direct = await asyncio.to_thread(extract_downloader_url, edit_url)
        payload, filename = await download_xlsx(direct)
        async with SessionLocal() as session:
            if kind == "schedule":
                items = parse_schedule_xlsx(payload, filename)
                count = await upsert_schedule(items, session)
            else:
                students, grades = parse_grades_xlsx(payload)
                count = await upsert_grades(students, grades, session)
        logger.info(f"[{name}] синхронизировано записей: {count}")
    except Exception as e:
        logger.error(f"[{name}] ошибка синхронизации: {e}")


async def sync_all() -> None:
    """Точка входа планировщика: расписание + успеваемость, каждые 2 дня."""
    await sync_source("расписание", SCHEDULE_EDIT_URL, "schedule")
    await sync_source("успеваемость", GRADES_EDIT_URL, "grades")
