"""Unit-тесты парсера Яндекс.Таблиц (реальный формат УЦП-25)."""

import datetime
import io

import pytest
from openpyxl import Workbook

from app.sync import (
    _file_group,
    _sheet_group,
    clean_subject_title,
    norm_name,
    parse_date,
    parse_lesson_cell,
    parse_time_range,
    parse_weekday,
    split_fio,
    student_code,
)


@pytest.mark.parametrize("raw,expected", [
    ("Пн", 1), ("понедельник", 1), ("1", 1), ("Вт", 2), ("среду", 3),
    ("Чт", 4), ("пятницу", 5), ("Сб", 6), ("Вс", 7), ("воскресенье", 7),
])
def test_parse_weekday(raw, expected):
    assert parse_weekday(raw) == expected


def test_parse_weekday_invalid():
    with pytest.raises(ValueError):
        parse_weekday("Фундимент")


def test_parse_time_range():
    assert parse_time_range("19:00-20:20") == ("19:00", "20:20")
    assert parse_time_range("11:00 - 12:20") == ("11:00", "12:20")
    assert parse_time_range("Сессия") is None
    assert parse_time_range("") is None


def test_parse_date():
    assert parse_date(datetime.datetime(2026, 1, 12)) == datetime.date(2026, 1, 12)
    assert parse_date(datetime.date(2026, 1, 12)) == datetime.date(2026, 1, 12)
    assert parse_date("01.09.2026") == datetime.date(2026, 9, 1)
    assert parse_date("07.06;2026") == datetime.date(2026, 6, 7)  # опечатка в таблице
    assert parse_date("2026-09-01") == datetime.date(2026, 9, 1)
    assert parse_date("") is None
    assert parse_date("не дата") is None


def test_split_fio_and_code():
    assert split_fio("Азимов Мурад Анар Оглы") == (
        "Азимов", "Мурад", "Азимов Мурад Анар Оглы")
    assert split_fio("Бутяков Максим ") == ("Бутяков", "Максим", "Бутяков Максим")
    # Код одинаков с отчеством и без — ведомости пишут по-разному.
    assert student_code("Аксёнов", "Роман") == norm_name("Аксёнов Роман")


def test_clean_subject_title():
    assert clean_subject_title(
        "Управление проектами\n (базовый уровень)\nРАНХиГС"
    ) == "Управление проектами (базовый уровень)"
    assert clean_subject_title(
        "Искусственный интеллект в менеджменте\nРАНХиГС\nДжимшер Челидзе"
    ) == "Искусственный интеллект в менеджменте"
    assert clean_subject_title(
        "Финансовый менеджмент и корпоративные финансы (зачет с оценкой)\nРАНХиГС"
    ) == "Финансовый менеджмент и корпоративные финансы"
    from app.sync import _is_formula_text, split_subject_title
    assert split_subject_title(
        "Финансовый менеджмент (зачет с оценкой)\nРАНХиГС"
    ) == ("Финансовый менеджмент", "зачет с оценкой")
    assert _is_formula_text("ИТОГОВАЯ ОЦЕНКА (зачет) = Тесты (60) + ...")
    assert _is_formula_text("Кт 1+ КТ 2+ защита = 100 баллов")
    assert _is_formula_text("УСПЕВАЕМОСТЬ \n1 СЕМЕСТР")
    assert not _is_formula_text("Управление командой\n(Нетология)")
    # Организация в скобках внутри названия — вырезать, название оставить.
    assert clean_subject_title(
        "Технико-экономическое обоснование проекта (Нетология)\n\nОксана Малервейн"
    ) == "Технико-экономическое обоснование проекта"


def test_lesson_cell_normal():
    r = parse_lesson_cell(
        "Финансовый менеджмент и корпоративные финансы\n"
        "РАНХиГС\n1\nАполлонов Александр Владимирович")
    assert r == {"kind": "lesson", "subject": "Финансовый менеджмент и корпоративные финансы",
                 "teacher": "Аполлонов Александр Владимирович", "org": "РАНХиГС",
                 "lesson_no": 1, "status": "active", "note": "", "link": ""}


def test_lesson_cell_link():
    r = parse_lesson_cell(
        "ДЕДЛАЙН\nДЗ 1 - команды здесь https://clck.ru/3TLESz\nМиМ")
    assert r["kind"] == "deadline"
    assert r["link"] == "https://clck.ru/3TLESz"
    r = parse_lesson_cell(
        "Финансовый менеджмент\nРАНХиГС\n1\nАполлонов Александр Владимирович")
    assert r["link"] == ""


def test_lesson_cell_cancelled_and_moved():
    r = parse_lesson_cell("ОТМЕНА\nМетоды исследований в менеджменте\n"
                          "РАНХиГС\n3\nВолков Андрей Иванович")
    assert r["status"] == "cancelled" and r["kind"] == "lesson"
    assert r["subject"] == "Методы исследований в менеджменте"
    r = parse_lesson_cell("ПЕРЕНОС на 29.03\nНИР\nРАНХиГС\nДобрусин Алексей\n1")
    assert r["status"] == "moved" and r["subject"] == "НИР"


def test_lesson_cell_attestation_event_deadline_marker():
    r = parse_lesson_cell("зачет, асинхрон\nМаркетинг\nГладышева Ксения")
    assert r["kind"] == "attestation" and r["subject"] == "Маркетинг"
    r = parse_lesson_cell("19:00 мск \nвстреча с представителем \nОзон финтех")
    assert r["kind"] == "event"
    r = parse_lesson_cell("ДЕДЛАЙН \nМиМ\nтесты 1-7 \nСДО до 23:59 мск")
    assert r["kind"] == "deadline"
    assert parse_lesson_cell("Сессия") is None
    assert parse_lesson_cell("КАНИКУЛЫ") is None
    assert parse_lesson_cell("") is None
    assert parse_lesson_cell("ОТМЕНА") is None  # маркер без предмета — не пара


def test_sheet_group_and_file_group():
    assert _sheet_group("2 семестр ", "УЦП-25") == "УЦП-25"
    assert _sheet_group("ЛФТ-25 2 сем", "УЦП-25") is None  # чужая группа
    assert _file_group("Расписание УЦП-25 1 сем 2025-2026.xlsx", "Расписание") == "УЦП-25"


def _schedule_wb():
    """Минимальный лист в реальном формате: шапка + неделя + пары."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2 семестр"
    ws["B1"] = "понедельник"
    ws["D1"] = "вторник"
    ws.merge_cells("B1:C1")
    ws["B2"] = "19:00-20:20"
    ws["C2"] = "20:40-22:00"
    ws["D2"] = "19:00-20:20"
    ws["E2"] = "20:40-22:00"
    ws["A3"] = 5
    ws["B3"] = datetime.datetime(2026, 2, 9)
    ws["D3"] = datetime.datetime(2026, 2, 10)
    ws["B4"] = "Финансовый менеджмент\nРАНХиГС\n1\nАполлонов Александр Владимирович"
    ws["D4"] = "ОТМЕНА\nМетоды исследований\nРАНХиГС\n1\nВолков Андрей Иванович"
    ws["B5"] = "Сессия"
    junk = wb.create_sheet("ЛФТ-25 2 сем")
    junk["B2"] = "19:00-20:20"
    junk["B3"] = datetime.datetime(2026, 2, 9)
    junk["B4"] = "Чужая пара\nРАНХиГС\n1\nНикто Никто"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_schedule_minisheet():
    from app.sync import parse_schedule_xlsx
    items = parse_schedule_xlsx(_schedule_wb(), "Расписание УЦП-25 1 сем.xlsx")
    assert len(items) == 2
    first, second = items
    assert first["group"] == "УЦП-25"
    assert first["date"] == "2026-02-09"
    assert (first["time_start"], first["time_end"]) == ("19:00", "20:20")
    assert first["status"] == "active" and first["lesson_no"] == 1
    assert second["date"] == "2026-02-10"
    assert second["status"] == "cancelled"


def _grades_wb():
    """Минимальная ведомость: шапка блока + задания + 2 студента."""
    wb = Workbook()
    ws = wb.active
    ws.title = "1 семестр 25-26"
    ws["F5"] = "Зачет с оценкой"
    ws["G5"] = "Управление проектами\nРАНХиГС"
    ws.merge_cells("G5:K5")
    headers = ["ФИО", "ДЗ-1", "Всего баллов", "Итоговая оценка", "словесн", "ECTS"]
    for i, h in enumerate(headers, start=7):
        ws.cell(7, i, h)
    ws.cell(10, 2, "Азимов Мурад Анар Оглы")
    ws.cell(10, 7, "Азимов Мурад")
    ws.cell(10, 8, 9)
    ws.cell(10, 9, 85)
    ws.cell(10, 10, 5)
    ws.cell(10, 11, "Отлично")
    ws.cell(10, 12, "B")
    ws.cell(11, 2, "Аксёнов Роман Николаевич")
    ws.cell(11, 7, "Аксёнов Роман")
    ws.cell(11, 10, "зачтено")
    ws.cell(12, 2, "Белкин Григорий Александрович")
    ws.cell(12, 7, "Белкин Григорий")
    ws.cell(12, 10, 4)
    # Второй блок: якорь-баллы + пустая шапка итоговой + ECTS над словесной.
    ws["L5"] = "Зачет с оценкой"
    ws["M5"] = "Маркетинг\nРАНХиГС"
    ws.merge_cells("M5:R5")
    for i, h in enumerate(["ДЗ", "Итого", "", "ECTS", ""], start=14):
        ws.cell(7, i, h)
    ws.cell(10, 15, 88)
    ws.cell(10, 16, 5)
    ws.cell(10, 17, "Отлично")
    ws.cell(10, 18, "B")
    ws.cell(11, 15, 70)
    ws.cell(11, 16, 4)
    ws.cell(11, 17, "Хорошо")
    ws.cell(11, 18, "C")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_grades_minisheet():
    from app.sync import _match_student, parse_grades_xlsx
    students, grades = parse_grades_xlsx(_grades_wb())
    assert {s["code"] for s in students} == {
        norm_name("Азимов Мурад"), norm_name("Аксёнов Роман"),
        norm_name("Белкин Григорий")}
    by_code = {g["student_code"]: g for g in grades
                 if g["subject"] == "Управление проектами"}
    assert len(grades) == 6  # +1 минимальная за пустую ячейку (Белкин/Маркетинг)
    g = by_code[norm_name("Азимов Мурад")]
    assert g["subject"] == "Управление проектами"
    assert g["attestation"] == "зачет с оценкой"
    assert (g["value"], g["verbal"], g["ects"], g["score"]) == ("5", "Отлично", "B", 85.0)
    assert by_code[norm_name("Аксёнов Роман")]["value"] == "зачтено"
    # Хитрый блок: баллы -> пустая шапка -> ECTS над словесной.
    mkt = [g for g in grades if g["subject"] == "Маркетинг"]
    assert len(mkt) == 3
    az = next(g for g in mkt if g["student_code"] == norm_name("Азимов Мурад"))
    assert (az["value"], az["verbal"], az["ects"], az["score"]) == (
        "5", "Отлично", "B", 88.0)
    assert az["attestation"] == "зачет с оценкой"
    belkin_min = next(g for g in grades
                      if g["subject"] == "Маркетинг"
                      and g["student_code"] == norm_name("Белкин Григорий"))
    assert (belkin_min["value"], belkin_min["verbal"],
            belkin_min["ects"], belkin_min["score"]) == (
        "2", "Неудовлетворительно", "F", 0.0)


def test_match_student_swapped_and_fuzzy():
    from app.sync import _match_student
    roster = {"aksenov": {"full_name": "Аксёнов Роман Николаевич"}}
    assert _match_student("Аксёнов Роман", roster) == "aksenov"
    assert _match_student("Роман Аксёнов", roster) == "aksenov"  # перевёрнуто
    assert _match_student("Аксенов Роман", roster) == "aksenov"  # е/ё
    assert _match_student("Неизвестный Человек", roster) is None
